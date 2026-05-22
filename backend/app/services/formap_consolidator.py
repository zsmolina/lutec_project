"""Consolidación de libro Excel Formap a reporte unificado estilizado."""

from __future__ import annotations

from io import BytesIO

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.exceptions import FormapConsolidationError

METADATA_KEYS = (
    "Entidad",
    "Departamento",
    "Municipio",
    "Dirección",
    "Tipo",
    "Horarios de operación",
)

COMPONENT_SHEETS = (
    "AIRES ACONDICIONADOS",
    "VENTILACIÓN",
    "REFRIGERACIÓN",
    "ILUMINACIÓN",
    "OFIMÁTICOS",
    "AIRE COMPRIMIDO",
    "SISTEMAS DE BOMBEO",
    "LUXOMETRÍA",
    "MEDICIÓN ELECTRICA-INSTANTANEA",
)

HEADER_KEYWORDS = (
    "ubicación",
    "ubicacion",
    "zona/área",
    "zona/area",
    "nombre\n(tablero - equipo)",
)

OUTPUT_SHEET_TITLE = "REPORTE UNIFICADO AE"
REPORT_TITLE = "INFORME CONSOLIDADO DE AUDITORÍA DE EFICIENCIA ENERGÉTICA"

COLOR_HEADER_PRINCIPAL = "2C4A6F"
COLOR_HEADER_TABLA = "4A6585"
COLOR_CEBRA = "F4F7FA"
COLOR_BORDE = "D3D3D3"


def _build_styles():
    fuente_titulo_principal = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    fuente_subtitulo_seccion = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fuente_cabecera_tabla = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
    fuente_datos = Font(name="Segoe UI", size=9.5, bold=False, color="000000")
    fuente_meta_clave = Font(name="Segoe UI", size=9.5, bold=True, color="2C4A6F")

    fill_titulo_principal = PatternFill(
        start_color=COLOR_HEADER_PRINCIPAL,
        end_color=COLOR_HEADER_PRINCIPAL,
        fill_type="solid",
    )
    fill_subtitulo = PatternFill(
        start_color=COLOR_HEADER_TABLA,
        end_color=COLOR_HEADER_TABLA,
        fill_type="solid",
    )
    fill_cebra = PatternFill(start_color=COLOR_CEBRA, end_color=COLOR_CEBRA, fill_type="solid")
    fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    borde_delgado = Border(
        left=Side(style="thin", color=COLOR_BORDE),
        right=Side(style="thin", color=COLOR_BORDE),
        top=Side(style="thin", color=COLOR_BORDE),
        bottom=Side(style="thin", color=COLOR_BORDE),
    )

    return {
        "titulo": fuente_titulo_principal,
        "subtitulo": fuente_subtitulo_seccion,
        "cabecera": fuente_cabecera_tabla,
        "datos": fuente_datos,
        "meta_clave": fuente_meta_clave,
        "fill_titulo": fill_titulo_principal,
        "fill_subtitulo": fill_subtitulo,
        "fill_cebra": fill_cebra,
        "fill_blanco": fill_blanco,
        "borde": borde_delgado,
    }


def _extract_metadata(excel_file: pd.ExcelFile) -> list[tuple[str, str]]:
    if "GENERAL" not in excel_file.sheet_names:
        return []

    df_general = pd.read_excel(excel_file, sheet_name="GENERAL", header=None)
    metadatos: list[tuple[str, str]] = []

    for _, fila in df_general.iterrows():
        valores_fila = [str(celda).strip() for celda in fila.values if pd.notna(celda)]
        for clave in METADATA_KEYS:
            if clave in valores_fila:
                idx_clave = valores_fila.index(clave)
                if idx_clave + 1 < len(valores_fila):
                    metadatos.append((clave, valores_fila[idx_clave + 1]))

    return metadatos


def _find_header_row(df_crudo: pd.DataFrame) -> int | None:
    for idx, fila in df_crudo.iterrows():
        valores_str = [str(c).lower() for c in fila.values if pd.notna(c)]
        if any(kw in valores_str for kw in HEADER_KEYWORDS):
            return int(idx)
    return None


def _read_component_sheet(excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame | None:
    df_crudo = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
    fila_encabezado = _find_header_row(df_crudo)
    if fila_encabezado is None:
        return None

    df_limpio = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=fila_encabezado)
    df_limpio = df_limpio.loc[:, df_limpio.columns.notna()]
    df_limpio.columns = [str(col).strip().replace("\n", " ") for col in df_limpio.columns]
    return df_limpio


def _autofit_columns(ws, max_columnas_global: int, fila_actual: int) -> None:
    for col in range(2, max_columnas_global + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, fila_actual):
            celda = ws.cell(row=row, column=col)
            if celda.coordinate in ws.merged_cells:
                continue
            if celda.value is not None:
                max_len = max(max_len, len(str(celda.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def consolidate_formap_workbook(source_bytes: bytes) -> bytes:
    """Transforma un Excel Formap de entrada en el reporte unificado estilizado."""
    if not source_bytes:
        raise FormapConsolidationError("El archivo Excel está vacío.")

    try:
        excel_file = pd.ExcelFile(BytesIO(source_bytes))
    except Exception as exc:
        raise FormapConsolidationError(
            "No se pudo leer el archivo. Verifique que sea un Excel válido (.xlsx)."
        ) from exc

    hojas_disponibles = excel_file.sheet_names
    if not hojas_disponibles:
        raise FormapConsolidationError("El libro no contiene hojas.")

    styles = _build_styles()
    metadatos = _extract_metadata(excel_file)

    wb_nuevo = openpyxl.Workbook()
    ws = wb_nuevo.active
    ws.title = OUTPUT_SHEET_TITLE
    ws.views.sheetView[0].showGridLines = True

    fila_actual = 2
    max_columnas_global = 10

    ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual + 1, end_column=10)
    for r in range(fila_actual, fila_actual + 2):
        for c in range(2, 11):
            ws.cell(row=r, column=c).fill = styles["fill_titulo"]

    celda_titulo = ws.cell(row=fila_actual, column=2)
    celda_titulo.value = REPORT_TITLE
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
    celda_titulo.font = styles["titulo"]
    fila_actual += 3

    if metadatos:
        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=5)
        for c in range(2, 6):
            ws.cell(row=fila_actual, column=c).fill = styles["fill_subtitulo"]

        c_sec = ws.cell(row=fila_actual, column=2)
        c_sec.value = "DATOS GENERALES DE LA EDIFICACIÓN AUDITADA"
        c_sec.font = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
        c_sec.alignment = Alignment(vertical="center")
        fila_actual += 1

        for clave, valor in metadatos:
            c_clv = ws.cell(row=fila_actual, column=2, value=clave.upper())
            c_clv.font = styles["meta_clave"]
            c_clv.border = styles["borde"]
            c_clv.fill = styles["fill_cebra"]

            ws.merge_cells(start_row=fila_actual, start_column=3, end_row=fila_actual, end_column=5)
            for c in range(3, 6):
                ws.cell(row=fila_actual, column=c).border = styles["borde"]

            c_val = ws.cell(row=fila_actual, column=3, value=valor)
            c_val.font = styles["datos"]
            c_val.alignment = Alignment(horizontal="left", vertical="center")
            fila_actual += 1

        fila_actual += 2

    bloques_incorporados = 0

    for hoja in COMPONENT_SHEETS:
        if hoja not in hojas_disponibles:
            continue

        df_limpio = _read_component_sheet(excel_file, hoja)
        if df_limpio is None:
            continue

        bloques_incorporados += 1
        esta_vacia = df_limpio.empty
        num_columnas = len(df_limpio.columns) if not esta_vacia else 6
        end_col_idx = 2 + max(num_columnas, 5) - 1
        max_columnas_global = max(max_columnas_global, end_col_idx)

        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=end_col_idx)
        for c in range(2, end_col_idx + 1):
            ws.cell(row=fila_actual, column=c).fill = styles["fill_subtitulo"]

        c_comp = ws.cell(row=fila_actual, column=2)
        c_comp.value = f"COMPONENTE: {hoja}"
        c_comp.font = styles["subtitulo"]
        c_comp.alignment = Alignment(vertical="center")
        ws.row_dimensions[fila_actual].height = 24
        fila_actual += 1

        if esta_vacia:
            ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=end_col_idx)
            for c in range(2, end_col_idx + 1):
                cell_vacia = ws.cell(row=fila_actual, column=c)
                cell_vacia.border = styles["borde"]
                cell_vacia.fill = styles["fill_cebra"]
            c_msg = ws.cell(row=fila_actual, column=2)
            c_msg.value = " No se registraron elementos de este componente en la inspección de campo."
            c_msg.font = Font(name="Segoe UI", size=9.5, italic=True, color="555555")
            c_msg.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[fila_actual].height = 20
            fila_actual += 3
            continue

        for col_idx, nombre_columna in enumerate(df_limpio.columns, start=2):
            celda_h = ws.cell(row=fila_actual, column=col_idx, value=nombre_columna)
            celda_h.font = styles["cabecera"]
            celda_h.fill = PatternFill(start_color="5A738E", end_color="5A738E", fill_type="solid")
            celda_h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda_h.border = styles["borde"]

        ws.row_dimensions[fila_actual].height = 28
        fila_actual += 1

        for row_idx, fila_datos in df_limpio.iterrows():
            fill_fila_actual = styles["fill_cebra"] if row_idx % 2 == 1 else styles["fill_blanco"]

            for col_idx, nombre_columna in enumerate(df_limpio.columns, start=2):
                valor_celda = fila_datos[nombre_columna]
                if pd.isna(valor_celda):
                    valor_celda = "-"

                celda_d = ws.cell(row=fila_actual, column=col_idx, value=valor_celda)
                celda_d.font = styles["datos"]
                celda_d.fill = fill_fila_actual
                celda_d.border = styles["borde"]

                if isinstance(valor_celda, (int, float)):
                    celda_d.alignment = Alignment(horizontal="right", vertical="center")
                    if any(
                        x in nombre_columna.lower()
                        for x in ["consumo", "potencia", "medida", "retilap"]
                    ):
                        celda_d.number_format = "#,##0.00"
                else:
                    celda_d.alignment = Alignment(horizontal="left", vertical="center")

            ws.row_dimensions[fila_actual].height = 19
            fila_actual += 1

        fila_actual += 2

    if bloques_incorporados == 0 and not metadatos:
        raise FormapConsolidationError(
            "No se encontraron hojas de componentes reconocibles ni datos en GENERAL."
        )

    _autofit_columns(ws, max_columnas_global, fila_actual)

    output = BytesIO()
    wb_nuevo.save(output)
    return output.getvalue()

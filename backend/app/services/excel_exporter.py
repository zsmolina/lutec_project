from pathlib import Path
from typing import Literal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.domain.billing_period import billing_period_from_reading_date
from app.domain.models.invoice import FacturaEnergiaData

ColumnFormat = Literal[
    "texto",
    "moneda",
    "decimal_fijo",
    "decimal_dinamico",
    "entero_miles",
    "entero",
    "tarifa",
]

ColumnConfig = tuple[str, ColumnFormat]

SHEET_TITLE = "Consolidado Facturas"
HEADER_FILL = "153426"

def _thin_border() -> Border:
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


def _build_value_map(data: FacturaEnergiaData) -> dict[str, object]:
    mes, anio = billing_period_from_reading_date(data.fecha_lectura_actual)
    return {
        "NIC": data.nic,
        "Documento Equivalente": data.documento_equivalente,
        "Fecha Emisión": data.fecha_emision,
        "Titular del Pago": data.titular_pago,
        "Estrato": data.estrato_clasificacion,
        "Propiedad del Activo": data.propiedad_activo,
        "Total a Pagar": data.total_a_pagar,
        "Energía Mes ($)": data.energia_mes_valor,
        "Factor Multiplicador": data.factor_multiplicador,
        "Consumo Actual (kWh)": data.consumo_actual_kwh,
        "Tipo de Consumo Calculado": data.tipo_consumo_calculado,
        "Fecha Lectura Anterior": data.fecha_lectura_anterior,
        "Fecha Lectura Actual": data.fecha_lectura_actual,
        "Mes": mes,
        "Año": anio,
        "Días Consumo": data.dias_consumo,
        "Consumo Diario (kWh)": data.consumo_diario,
        "G (Generación)": data.tarifa_cu.generacion,
        "T (Transmisión)": data.tarifa_cu.transmision,
        "P (Pérdidas)": data.tarifa_cu.perdidas,
        "R (Restricciones)": data.tarifa_cu.restricciones,
        "D (Distribución)": data.tarifa_cu.distribucion,
        "C (Comercialización)": data.tarifa_cu.comercializacion,
        "CU Aplicado": data.tarifa_cu.cu_aplicado,
    }


def _apply_number_format(cell, fmt: ColumnFormat) -> None:
    formats: dict[ColumnFormat, str] = {
        "moneda": '"$"#,##0.00',
        "decimal_fijo": "#,##0.00",
        "decimal_dinamico": "0.00",
        "entero_miles": "#,##0",
        "entero": "#,##0",
        "tarifa": "0.00##",
    }
    if fmt in formats:
        cell.number_format = formats[fmt]


class ExcelExporter:
    """Consolida facturas extraídas en un libro Excel unificado."""

    def __init__(self, output_path: Path, column_config: list[ColumnConfig]):
        self.output_path = output_path
        self.column_config = column_config
        self._wb, self._ws = self._load_or_create()

    def _load_or_create(self):
        if self.output_path.exists():
            wb = openpyxl.load_workbook(self.output_path)
            return wb, wb.active

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_TITLE
        ws.views.sheetView[0].showGridLines = True

        fill_header = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        border = _thin_border()

        headers = [header for header, _ in self.column_config]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        ws.row_dimensions[1].height = 28
        return wb, ws

    def append_invoice(self, data: FacturaEnergiaData) -> None:
        font_data = Font(name="Segoe UI", size=11, color="000000")
        border = _thin_border()
        values = _build_value_map(data)

        row = self._ws.max_row + 1
        self._ws.row_dimensions[row].height = 22

        for col_idx, (header, fmt) in enumerate(self.column_config, 1):
            cell = self._ws.cell(row=row, column=col_idx, value=values[header])
            cell.font = font_data
            cell.border = border

            if fmt == "texto":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = Font(name="Consolas", size=10, color="000000")
                cell.alignment = Alignment(horizontal="right", vertical="center")
                _apply_number_format(cell, fmt)

    def finalize(self) -> None:
        self._autosize_columns()
        self.save()

    def _autosize_columns(self) -> None:
        for col in self._ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            letter = get_column_letter(col[0].column)
            self._ws.column_dimensions[letter].width = max(max_len + 4, 12)

    def save(self) -> None:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(self.output_path)
